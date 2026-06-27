import os
import bcrypt
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client
from datetime import date, timedelta

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
API_SECRET = os.environ.get("API_SECRET")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
app = FastAPI(title="Sistema de Equipos LAMYG")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Verificar API Key
@app.middleware("http")
async def verificar_api_key(request: Request, call_next):
    rutas_publicas = ["/", "/docs", "/openapi.json", "/redoc"]
    if request.url.path in rutas_publicas:
        return await call_next(request)
    api_key = request.headers.get("X-API-Key")
    if api_key != API_SECRET:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=401, content={"detail": "API Key inválida"})
    return await call_next(request)

@app.get("/")
def inicio():
    return {"mensaje": "API Equipos LAMYG activa"}

# ==================== AUTH ====================
@app.post("/auth/registro")
def registrar_usuario(datos: dict):
    email = datos.get("email", "").strip().lower()
    password = datos.get("password", "")
    nombre = datos.get("nombre", "").strip()
    codigo_lab = datos.get("codigo_laboratorio", "").strip()
    nombre_lab = datos.get("nombre_laboratorio", "").strip()

    if not email or not password or not nombre:
        raise HTTPException(status_code=400, detail="Email, contraseña y nombre son obligatorios")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")

    existe = supabase.table("usuarios").select("id").eq("email", email).execute()
    if existe.data:
        raise HTTPException(status_code=400, detail="Ya existe un usuario con ese email")

    if codigo_lab:
        lab = supabase.table("laboratorios").select("id").eq("codigo_acceso", codigo_lab).execute()
        if not lab.data:
            raise HTTPException(status_code=404, detail="Código de laboratorio no encontrado")
        lab_id = lab.data[0]["id"]
        rol = "usuario"
    elif nombre_lab:
        import random
        import string
        codigo_nuevo = nombre_lab[:4].upper() + "-" + ''.join(random.choices(string.digits, k=4))
        nuevo_lab = supabase.table("laboratorios").insert({
            "nombre": nombre_lab,
            "codigo_acceso": codigo_nuevo
        }).execute()
        lab_id = nuevo_lab.data[0]["id"]
        rol = "admin"
    else:
        raise HTTPException(status_code=400, detail="Debes ingresar un código de laboratorio o crear uno nuevo")

    password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    usuario = supabase.table("usuarios").insert({
        "email": email,
        "password_hash": password_hash,
        "nombre": nombre,
        "laboratorio_id": lab_id,
        "rol": rol,
    }).execute()

    return {
        "usuario_id": usuario.data[0]["id"],
        "nombre": nombre,
        "email": email,
        "rol": rol,
        "laboratorio_id": lab_id,
        "codigo_laboratorio": codigo_lab if codigo_lab else codigo_nuevo,
    }
@app.post("/auth/login")
def login(datos: dict):
    email = datos.get("email", "").strip().lower()
    password = datos.get("password", "")

    if not email or not password:
        raise HTTPException(status_code=400, detail="Email y contraseña son obligatorios")

    usuario = supabase.table("usuarios").select("*").eq("email", email).execute()
    if not usuario.data:
        raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")

    user = usuario.data[0]
    if not bcrypt.checkpw(password.encode('utf-8'), user["password_hash"].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")

    # Obtener datos del laboratorio
    lab = supabase.table("laboratorios").select("*").eq("id", user["laboratorio_id"]).execute()
    if not lab.data:
        raise HTTPException(status_code=404, detail="Laboratorio no encontrado")

    lab_data = lab.data[0]

    # Verificar si el laboratorio está activo
    if not lab_data.get("activo", True):
        raise HTTPException(status_code=403, detail="CUENTA_INACTIVA")

    # Verificar fecha de vencimiento
    fecha_venc = lab_data.get("fecha_vencimiento")
    if fecha_venc:
        from datetime import date as date_class
        if date_class.fromisoformat(fecha_venc) < date_class.today():
            raise HTTPException(status_code=403, detail="TRIAL_VENCIDO")

    dias_restantes = 0
    if fecha_venc:
        from datetime import date as date_class
        dias_restantes = (date_class.fromisoformat(fecha_venc) - date_class.today()).days

    return {
        "usuario_id": user["id"],
        "nombre": user["nombre"],
        "email": user["email"],
        "rol": user["rol"],
        "laboratorio_id": user["laboratorio_id"],
        "laboratorio_nombre": lab_data["nombre"],
        "codigo_laboratorio": lab_data["codigo_acceso"],
        "dias_restantes": dias_restantes,
    }

# ==================== EQUIPOS ====================

@app.get("/equipos")
def listar_equipos(lab_id: str = None):
    query = supabase.table("equipos").select("*").order("nombre")
    if lab_id:
        query = query.eq("laboratorio_id", lab_id)
    result = query.execute()
    return result.data

@app.get("/equipos/exportar")
def exportar_equipos(lab_id: str = None):
    import openpyxl
    from fastapi.responses import Response
    from io import BytesIO

    query = supabase.table("equipos").select("*").order("nombre")
    if lab_id:
        query = query.eq("laboratorio_id", lab_id)
    result = query.execute()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipos"

    # Encabezados
    headers = ["Nombre", "Código", "Marca", "N° Serie", "Rango/Capacidad",
               "Fecha Calibración", "Próxima Calibración", "Calibrado por",
               "Responsable", "Estado"]
    ws.append(headers)

    # Estilo encabezados
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")

    # Datos
    for e in result.data:
        ws.append([
            e.get("nombre"), e.get("codigo"), e.get("marca"),
            e.get("numero_serie"), e.get("rango_capacidad"),
            e.get("fecha_calibracion"), e.get("fecha_proxima_calibracion"),
            e.get("calibrado_por"), e.get("responsable"), e.get("estado"),
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=equipos.xlsx"}
    )

@app.get("/equipos/alertas")
def alertas_calibracion(lab_id: str = None):
    hoy = date.today()
    limite = hoy + timedelta(days=60)
    query = supabase.table("equipos").select(
        "nombre, codigo, fecha_proxima_calibracion, calibrado_por, responsable, laboratorio_id"
    ).lte("fecha_proxima_calibracion", str(limite)).order("fecha_proxima_calibracion")
    if lab_id:
        query = query.eq("laboratorio_id", lab_id)
    result = query.execute()

    equipos = []
    for e in result.data:
        fecha = date.fromisoformat(e["fecha_proxima_calibracion"])
        dias = (fecha - hoy).days
        e["dias_para_vencer"] = dias
        e["estado_alerta"] = "vencido" if dias < 0 else "critico" if dias <= 30 else "proximo"
        equipos.append(e)
    return equipos

@app.get("/equipos/{codigo}")
def detalle_equipo(codigo: str):
    result = supabase.table("equipos").select("*").eq("codigo", codigo).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    return result.data[0]

@app.post("/equipos")
def crear_equipo(equipo: dict):
    if not equipo.get("nombre") or not equipo.get("codigo"):
        raise HTTPException(status_code=400, detail="Nombre y código son obligatorios")
    existe = supabase.table("equipos").select("codigo").eq("codigo", equipo["codigo"]).execute()
    if existe.data:
        raise HTTPException(status_code=400, detail="Ya existe un equipo con ese código")
    campos_permitidos = [
        'nombre', 'codigo', 'marca', 'modelo', 'numero_serie',
        'rango_capacidad', 'resolucion', 'normas_asociadas', 'tipo',
        'fecha_verificacion', 'fecha_proxima_verificacion',
        'fecha_calibracion', 'fecha_proxima_calibracion',
        'periodo_calibracion', 'calibrado_por',
        'fecha_mantenimiento', 'fecha_proximo_mantenimiento',
        'periodo_mantenimiento', 'puntos_calibracion',
        'responsable', 'observaciones', 'estado', 'laboratorio_id'
    ]
    datos_limpios = {k: v for k, v in equipo.items() if k in campos_permitidos}
    result = supabase.table("equipos").insert(datos_limpios).execute()
    return result.data

@app.put("/equipos/{codigo}")
def actualizar_equipo(codigo: str, datos: dict):
    existe = supabase.table("equipos").select("codigo").eq("codigo", codigo).execute()
    if not existe.data:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    campos_permitidos = [
        'nombre', 'marca', 'modelo', 'numero_serie',
        'rango_capacidad', 'resolucion', 'normas_asociadas', 'tipo',
        'fecha_verificacion', 'fecha_proxima_verificacion',
        'fecha_calibracion', 'fecha_proxima_calibracion',
        'periodo_calibracion', 'calibrado_por',
        'fecha_mantenimiento', 'fecha_proximo_mantenimiento',
        'periodo_mantenimiento', 'puntos_calibracion',
        'responsable', 'observaciones', 'estado'
    ]
    datos_limpios = {k: v for k, v in datos.items() if k in campos_permitidos and v is not None}
    if not datos_limpios:
        raise HTTPException(status_code=400, detail="No hay datos válidos para actualizar")
    response = supabase.table("equipos").update(datos_limpios).eq("codigo", codigo).execute()
    return response.data[0]

@app.put("/equipos/{codigo}/calibracion")
def registrar_calibracion(codigo: str, datos: dict):
    existe = supabase.table("equipos").select("codigo").eq("codigo", codigo).execute()
    if not existe.data:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    result = supabase.table("equipos").update(datos).eq("codigo", codigo).execute()
    return result.data



@app.delete("/equipos/{codigo}")
def eliminar_equipo(codigo: str):
    existe = supabase.table("equipos").select("codigo").eq("codigo", codigo).execute()
    if not existe.data:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    supabase.table("equipos").delete().eq("codigo", codigo).execute()
    return {"mensaje": f"Equipo {codigo} eliminado"}